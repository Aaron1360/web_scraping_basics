"""This script shows how to retrieve the content of a web page using requests and Beautiful Soup.
The page we are going to work on is a sandbox of www.scrapethissite.com"""

import requests
from bs4 import BeautifulSoup as bs
import openpyxl as xl

def scrape_data(countries):
    """This function organizes the desired data on a list of dictionaries.

    Args:
        countries (bs4.element.ResultSet): Object that holds all the HTML tags that match the specified criteria.

    Returns:
        list: List of dictionaries where each dictionary represents a country.
    """ 
    list_of_countries = []
    for country in countries:
        new_country = {}
        new_country["Name"] = country.h3.text.strip()
        new_country["Capital"]= country.find("span",class_="country-capital").text.strip()
        new_country["Population"] = country.find("span",class_="country-population").text.strip()
        new_country["Area"] = country.find("span",class_="country-area").text.strip()
        list_of_countries.append(new_country)
    return list_of_countries

def show_info(data):
    for country in data:
        for key,value in country.items():
            print(f"{key}: {value}")
        print()

def save_txt(data):
    """Stores the information on list_of_countries.txt

    Args:
        data (list): List of dictionaries with country information.
    """
    with open("list_of_countries.txt","w") as file:
        for i,country in enumerate(data):
            file.write(f"{i+1}.- ")
            for key,value in country.items():
                if key == "Name":
                    file.write(f"{key}: {value}.\n")
                else:
                    file.write(f"\t{key}: {value}.\n")
            file.write("\n")
    print("**Data saved successfully**")

def save_excel(data):
    """Create excel file and organizes the scraped info.

    Args:
        data (list): List of dictionaries with country information.
    """
    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Countries"
    keys = data[0].keys()
    for col, key in enumerate(keys,1):
        ws.cell(1,col,key)
    for row in range(2,len(data)+2):
        for col,info in enumerate(data[row-2].values(),1):
            ws.cell(row,col,info)
    wb.save("list_of_countries.xlsx")
    
# Target
page = "https://www.scrapethissite.com/pages/simple/"

# Make request and parse content
response = requests.get(page)
# response.encoding = "latin-1"
soup = bs(response.content,"html.parser",from_encoding="latin-1")    
countries = soup.find_all(name="div",class_="col-md-4 country")

# Create files
data = scrape_data(countries)
save_txt(data)
save_excel(data)

